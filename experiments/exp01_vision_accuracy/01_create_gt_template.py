"""GT 라벨링 엑셀 템플릿 생성.

images/ 폴더의 100개 상품 이미지를 읽어서
구글 시트에서 협업 라벨링할 수 있는 엑셀을 만든다.
담당자 분배 포함, 첫 행에 예시 포함.

Usage:
    python 01_create_gt_template.py
"""

import sys
from pathlib import Path
from collections import defaultdict, Counter

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# 1. 이미지 목록 수집
# ---------------------------------------------------------------------------

IMAGES_DIR = ROOT / "images"
image_ids = sorted([d.name for d in IMAGES_DIR.iterdir() if d.is_dir()])
print(f"이미지 폴더 수: {len(image_ids)}")


def get_platform(image_id: str) -> str:
    prefix = image_id.split("_")[0]
    return {
        "ably": "에이블리",
        "musinsa": "무신사",
        "zigzag": "지그재그",
    }.get(prefix, "기타")


# ---------------------------------------------------------------------------
# 2. 담당자 분배 (플랫폼별 번갈아 균등 배분)
# ---------------------------------------------------------------------------

by_platform = defaultdict(list)
for img_id in image_ids:
    by_platform[get_platform(img_id)].append(img_id)

assign_num_meta = {}   # 숫자+메타 담당
assign_style_mkt = {}  # 스타일+마케팅 담당

counter_a = counter_b = 0
for platform in ["에이블리", "무신사", "지그재그", "기타"]:
    for img_id in by_platform[platform]:
        assign_num_meta[img_id] = "경현" if counter_a % 2 == 0 else "팜팜이"
        counter_a += 1
        assign_style_mkt[img_id] = "낭연" if counter_b % 2 == 0 else "정현"
        counter_b += 1

print(f"숫자+메타: {Counter(assign_num_meta.values())}")
print(f"스타일+마케팅: {Counter(assign_style_mkt.values())}")


# ---------------------------------------------------------------------------
# 3. 엑셀 생성
# ---------------------------------------------------------------------------

wb = Workbook()
ws = wb.active
ws.title = "GT 라벨링"

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 그룹별 색상
FILL_INFO = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_NUM = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
FILL_META = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
FILL_STYLE = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
FILL_MKT = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_OPT = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
FILL_NOTE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_EXAMPLE = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# 열 정의: (이름, 그룹색, 너비)
COLUMNS = [
    # 기본정보
    ("image_id", FILL_INFO, 22),
    ("platform", FILL_INFO, 10),
    ("image_path", FILL_INFO, 30),
    ("담당_숫자메타", FILL_INFO, 14),
    ("담당_스타일마케팅", FILL_INFO, 16),
    # 텍스트 추출 (숫자)
    ("product_name", FILL_NUM, 30),
    ("original_price", FILL_NUM, 14),
    ("has_discount(0/1)", FILL_NUM, 16),
    ("discounted_price", FILL_NUM, 16),
    ("discount_rate(%)", FILL_NUM, 14),
    ("review_count", FILL_NUM, 13),
    ("review_score", FILL_NUM, 13),
    ("wishlist_count", FILL_NUM, 14),
    # 메타
    ("shot_type", FILL_META, 16),
    ("visibility", FILL_META, 12),
    # 스타일 (최대 3개)
    ("style_keyword_1", FILL_STYLE, 14),
    ("style_keyword_2", FILL_STYLE, 14),
    ("style_keyword_3", FILL_STYLE, 14),
    # 마케팅 트리거
    ("trend_hype(0/1)", FILL_MKT, 15),
    ("bundle(0/1)", FILL_MKT, 12),
    ("confidence(0/1)", FILL_MKT, 15),
    ("marketing_phrases", FILL_MKT, 30),
    # 선택 필드
    ("delivery_info", FILL_OPT, 20),
    ("brand", FILL_OPT, 16),
    # 비고
    ("비고", FILL_NOTE, 25),
]

# 그룹 헤더 (행 1)
GROUP_SPANS = [
    ("기본정보", 1, 5, FILL_INFO),
    ("텍스트 추출 (숫자)", 6, 13, FILL_NUM),
    ("메타", 14, 15, FILL_META),
    ("스타일 (최대 3개)", 16, 18, FILL_STYLE),
    ("마케팅 트리거", 19, 22, FILL_MKT),
    ("선택 필드", 23, 24, FILL_OPT),
    ("", 25, 25, FILL_NOTE),
]

for group_name, start_col, end_col, fill in GROUP_SPANS:
    if start_col != end_col:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=1, column=c, value=group_name if c == start_col else None)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

# 필드 헤더 (행 2)
for col_idx, (col_name, fill, width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=col_name)
    cell.fill = fill
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 4. 예시 행 (행 3)
# ---------------------------------------------------------------------------

first_id = image_ids[0]
EXAMPLE_ROW = [
    first_id,                       # image_id
    get_platform(first_id),         # platform
    f"images/{first_id}/",          # image_path
    assign_num_meta[first_id],      # 담당_숫자메타
    assign_style_mkt[first_id],     # 담당_스타일마케팅
    "[밍크안감] 리버시블 항공 점퍼", # product_name
    59900,                          # original_price
    1,                              # has_discount
    53910,                          # discounted_price
    10,                             # discount_rate
    4826,                           # review_count
    4.8,                            # review_score
    92000,                          # wishlist_count
    "모델 착용샷",                  # shot_type
    "양호",                         # visibility
    "캐주얼",                       # style_keyword_1
    "심플베이직",                    # style_keyword_2
    None,                           # style_keyword_3
    1,                              # trend_hype
    0,                              # bundle
    1,                              # confidence
    "인기폭발, 누적판매 5만",       # marketing_phrases
    "무료배송",                     # delivery_info
    "스파오",                       # brand
    "← 예시 행 (수정하세요)",       # 비고
]

for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
    cell = ws.cell(row=3, column=col_idx, value=value)
    cell.border = THIN_BORDER
    cell.fill = FILL_EXAMPLE
    cell.alignment = Alignment(vertical="center")


# ---------------------------------------------------------------------------
# 5. 나머지 이미지 행 (행 4~)
# ---------------------------------------------------------------------------

for row_idx, img_id in enumerate(image_ids[1:], start=4):
    row_data = [
        img_id,
        get_platform(img_id),
        f"images/{img_id}/",
        assign_num_meta[img_id],
        assign_style_mkt[img_id],
    ]
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
    # 나머지 빈 칸에도 테두리
    for col_idx in range(len(row_data) + 1, len(COLUMNS) + 1):
        ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER


# ---------------------------------------------------------------------------
# 6. 드롭다운
# ---------------------------------------------------------------------------

STYLE_KEYWORDS = [
    "심플베이직", "락시크", "힙", "페미닌", "러블리",
    "모리걸", "빈티지", "스트릿", "캐주얼", "섹시글램",
]
SHOT_TYPES = ["모델 착용샷", "흰 배경 단독샷", "행거샷", "기타"]
VISIBILITY = ["양호", "부분가림", "불량"]

last_row = len(image_ids) + 2  # 헤더2 + 데이터100

# style_keywords (열 16, 17, 18)
dv_style = DataValidation(
    type="list",
    formula1=f'"{",".join(STYLE_KEYWORDS)}"',
    allow_blank=True,
)
dv_style.error = "10종 스타일 키워드 중 선택하세요"
ws.add_data_validation(dv_style)
for col in [16, 17, 18]:
    dv_style.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last_row}")

# shot_type (열 14)
dv_shot = DataValidation(
    type="list",
    formula1=f'"{",".join(SHOT_TYPES)}"',
    allow_blank=True,
)
ws.add_data_validation(dv_shot)
dv_shot.add(f"N3:N{last_row}")

# visibility (열 15)
dv_vis = DataValidation(
    type="list",
    formula1=f'"{",".join(VISIBILITY)}"',
    allow_blank=True,
)
ws.add_data_validation(dv_vis)
dv_vis.add(f"O3:O{last_row}")

# 0/1 드롭다운 (has_discount=8, trend_hype=19, bundle=20, confidence=21)
dv_binary = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
ws.add_data_validation(dv_binary)
for col in [8, 19, 20, 21]:
    dv_binary.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last_row}")

# 헤더 고정 + 필터
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{last_row}"


# ---------------------------------------------------------------------------
# 7. 참조 시트: 라벨링 가이드
# ---------------------------------------------------------------------------

ws_guide = wb.create_sheet("라벨링 가이드")
ws_guide.column_dimensions["A"].width = 22
ws_guide.column_dimensions["B"].width = 65

guide = [
    ("필드", "설명"),
    ("", ""),
    ("[ 텍스트 추출 ]", "이미지에서 보이는 값을 그대로 입력. 안 보이면 비워두기"),
    ("product_name", "상품명 전체 (마케팅 문구 포함). 예: [밍크안감] 리버시블 항공 점퍼"),
    ("original_price", "원가 (숫자만). 예: 59900"),
    ("has_discount(0/1)", "할인 여부. 할인 표시 있으면 1, 없으면 0"),
    ("discounted_price", "할인가 (숫자만). 예: 53910"),
    ("discount_rate(%)", "할인율 (숫자만). 예: 10"),
    ("review_count", "리뷰 수. 예: 4826"),
    ("review_score", "평점 (0.0~5.0). 예: 4.8"),
    ("wishlist_count", "좋아요/찜 수. 예: 92000"),
    ("", ""),
    ("[ 메타 ]", ""),
    ("shot_type", "모델 착용샷 / 흰 배경 단독샷 / 행거샷 / 기타"),
    ("visibility", "양호(전체 선명) / 부분가림(일부 가림) / 불량(대부분 안보임)"),
    ("", ""),
    ("[ 스타일 ]", "10종 중 최대 3개 선택 (드롭다운)"),
    ("10종 목록", "심플베이직, 락시크, 힙, 페미닌, 러블리, 모리걸, 빈티지, 스트릿, 캐주얼, 섹시글램"),
    ("", ""),
    ("[ 마케팅 트리거 ]", "상품명/이미지에서 마케팅 문구 유무 판단"),
    ("trend_hype(0/1)", "유행 자극 문구. 예: 인기폭발, 품절임박, HOT, 대란템"),
    ("bundle(0/1)", "묶음/세트 문구. 예: 1+1, 세트특가, 사은품증정"),
    ("confidence(0/1)", "신뢰/보증 문구. 예: 누적판매 10만, 만족도 99%"),
    ("marketing_phrases", "해당 마케팅 문구 원문 (쉼표 구분)"),
    ("", ""),
    ("[ 선택 필드 ]", "이미지에서 보이면 입력, 안 보이면 비워두기"),
    ("delivery_info", "배송 정보. 예: 무료배송, 당일발송, 빠른출발"),
    ("brand", "브랜드명. 예: 스파오, 무탠다드"),
    ("", ""),
    ("[ 주의사항 ]", ""),
    ("", "- 안 보이는 필드는 비워두기 (추측 X)"),
    ("", "- 쿠폰 할인 ≠ 기본 할인 (쿠폰가는 무시)"),
    ("", "- 좋아요 ≠ 브랜드 팔로워 (혼동 주의)"),
    ("", "- 지그재그는 좋아요 없음 → wishlist_count 비워두기"),
]

for row_idx, (field, desc) in enumerate(guide, start=1):
    a = ws_guide.cell(row=row_idx, column=1, value=field)
    b = ws_guide.cell(row=row_idx, column=2, value=desc)
    if field.startswith("["):
        a.font = Font(bold=True, size=11)
    elif row_idx == 1:
        a.font = Font(bold=True)
        b.font = Font(bold=True)


# ---------------------------------------------------------------------------
# 8. 저장
# ---------------------------------------------------------------------------

OUTPUT_DIR = Path(__file__).resolve().parent / "gt"
OUTPUT_DIR.mkdir(exist_ok=True)

# 기존 파일 정리
for old in OUTPUT_DIR.glob("*.xlsx"):
    old.unlink()
    print(f"삭제: {old.name}")

output_path = OUTPUT_DIR / "gt_labeling_v1_assigned.xlsx"
wb.save(output_path)

print(f"\n저장: {output_path}")
print(f"이미지 {len(image_ids)}개, 열 {len(COLUMNS)}개, 예시 1행 포함")
