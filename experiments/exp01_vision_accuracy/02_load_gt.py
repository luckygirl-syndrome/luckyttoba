"""GT 엑셀 → JSONL 변환.

Usage:
    python 02_load_gt.py [--input gt/gt_labels_filled.xlsx] [--output gt/gt_labels.jsonl]

엑셀 헤더(2행) 기준으로 각 행을 1개 JSON 객체로 변환.
- style_keyword_1/2/3 → style_keywords: [...] (None 제외)
- 단위 suffix (0/1), (%) 제거
- 빈 셀은 null
- 일관성 검증 경고 출력
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


HEADER_RENAME = {
    "has_discount(0/1)": "has_discount",
    "discount_rate(%)": "discount_rate",
    "trend_hype(0/1)": "trend_hype",
    "bundle(0/1)": "bundle",
    "confidence(0/1)": "confidence",
}

# JSONL에 그대로 키로 보낼 필드 (style_keyword_*, 메타 컬럼 제외)
# marketing_phrases는 새 GT 엑셀에서 빠짐 → 빈 list로 채움 (evaluate에서 skip 처리)
# delivery_fee는 새로 추가됨 (배송 비용 텍스트)
KEEP_FIELDS = {
    "image_id", "platform",
    "product_name", "original_price", "has_discount", "discounted_price",
    "discount_rate", "review_count", "review_score", "wishlist_count",
    "shot_type", "visibility",
    "trend_hype", "bundle", "confidence",
    "delivery_info", "delivery_fee", "brand",
}

NUMERIC_INT = {"original_price", "discounted_price", "discount_rate",
               "review_count", "wishlist_count",
               "has_discount", "trend_hype", "bundle", "confidence"}
NUMERIC_FLOAT = {"review_score"}


def normalize_cell(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


_NUM_RE = __import__("re").compile(r"-?\d+(?:\.\d+)?")


def to_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").strip()
    if not s:
        return None
    m = _NUM_RE.search(s)  # "9999+", "999개" 같은 표기 허용
    return int(float(m.group())) if m else None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s:
        return None
    m = _NUM_RE.search(s)
    return float(m.group()) if m else None


def parse_args():
    ap = argparse.ArgumentParser()
    base = Path(__file__).resolve().parent
    ap.add_argument("--input", default=str(base / "gt" / "gt_labels_filled.xlsx"),
                    help="GT 엑셀 경로 (기본: gt/gt_labels_filled.xlsx)")
    ap.add_argument("--output", default=str(base / "gt" / "gt_labels.jsonl"),
                    help="출력 JSONL 경로")
    ap.add_argument("--sheet", default="GT 라벨링")
    ap.add_argument("--skip-example", action="store_true", default=True,
                    help="3행(예시 행) 스킵 (기본 on)")
    ap.add_argument("--include-example", dest="skip_example", action="store_false")
    return ap.parse_args()


def main():
    args = parse_args()
    in_path = Path(args.input)
    out_path = Path(args.output)

    if not in_path.exists():
        sys.exit(f"입력 엑셀이 없습니다: {in_path}\n"
                 f"라벨링이 끝난 파일을 {in_path.name}로 저장하거나 --input 으로 경로를 지정하세요.")

    wb = load_workbook(in_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        sys.exit(f"시트 '{args.sheet}'를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}")
    ws = wb[args.sheet]

    raw_headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [HEADER_RENAME.get(h, h) for h in raw_headers]

    start_row = 4 if args.skip_example else 3

    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0
    warnings = []

    with out_path.open("w", encoding="utf-8") as f:
        for r in range(start_row, ws.max_row + 1):
            cells = {headers[c - 1]: normalize_cell(ws.cell(row=r, column=c).value)
                     for c in range(1, ws.max_column + 1)}

            if not cells.get("image_id"):
                continue  # 빈 행

            # style_keywords 합치기
            sk = [cells.get(f"style_keyword_{i}") for i in (1, 2, 3)]
            sk = [x for x in sk if x]

            # marketing_phrases: 새 GT 엑셀은 marketing_phrases 컬럼이 없고
            # "비고" / "비고 (애매한 건 주황)" 컬럼에 마케팅 문구를 적어둠.
            # 우선순위: marketing_phrases 컬럼 > 비고 컬럼.
            # 구분자는 "/", ",", "·" 모두 허용.
            mp_raw = cells.get("marketing_phrases")
            if not mp_raw:
                # 비고 컬럼 찾기 (헤더 이름이 "비고"로 시작하는 모든 컬럼)
                for h, v in cells.items():
                    if h and isinstance(h, str) and h.strip().startswith("비고") and v:
                        mp_raw = v
                        break
            if mp_raw:
                # 여러 구분자 split
                import re as _re
                tokens = _re.split(r"[/,·、|]+", str(mp_raw))
                mp = [t.strip() for t in tokens if t and t.strip()]
            else:
                mp = []

            record = {}
            for k in KEEP_FIELDS:
                v = cells.get(k)
                if k in NUMERIC_INT:
                    record[k] = to_int(v)
                elif k in NUMERIC_FLOAT:
                    record[k] = to_float(v)
                else:
                    record[k] = v
            record["style_keywords"] = sk
            # GT에 marketing_phrases 컬럼이 빠졌어도 호환성 위해 기록 (비고 활용 가능)
            record["marketing_phrases"] = mp

            # ---- 일관성 검증 ----
            img_id = record["image_id"]
            if record.get("has_discount") == 0 and record.get("discounted_price") is not None:
                warnings.append(f"{img_id}: has_discount=0인데 discounted_price={record['discounted_price']}")
            if record.get("has_discount") == 1 and record.get("discounted_price") is None and record.get("discount_rate") is None:
                warnings.append(f"{img_id}: has_discount=1인데 discounted_price/discount_rate 둘 다 null")
            if img_id.startswith("zigzag") and record.get("wishlist_count") is not None:
                warnings.append(f"{img_id}: 지그재그인데 wishlist_count={record['wishlist_count']} (지그재그는 좋아요 없음 → null이어야)")

            f.write(json.dumps(record, ensure_ascii=False) + "\n")
            rows_written += 1

    print(f"[OK] {rows_written} rows → {out_path}")
    if warnings:
        print(f"\n경고 {len(warnings)}건:")
        for w in warnings[:30]:
            print(f"  - {w}")
        if len(warnings) > 30:
            print(f"  ... 외 {len(warnings) - 30}건")


if __name__ == "__main__":
    main()
