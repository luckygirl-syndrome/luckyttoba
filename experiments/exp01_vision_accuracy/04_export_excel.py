"""평가 결과를 보기 좋은 엑셀로 export.

Inputs:
  outputs/eval_results.parquet
  outputs/summary.json
  gt/gt_labels.jsonl
  ../exp00_vision_extraction/results/gemini_v2/*.jsonl

Output:
  outputs/score_summary.xlsx — 시트 5개:
    1. 요약 (필드별 정확도 + 색상)
    2. 플랫폼별 (플랫폼 × 필드 히트맵 — 색상 그라데이션)
    3. 마케팅 P-R-F1 (3축 + confusion matrix)
    4. 플래그 (follower_confusion 27건, coupon_confusion 등)
    5. 이미지별 상세 (image_id × 필드, status 컬러)
"""

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
OUT_DIR = BASE / "outputs"
GT_PATH = BASE / "gt" / "gt_labels.jsonl"
PRED_DIR = BASE.parent / "exp00_vision_extraction" / "results" / "gemini_v3"

# ---------- 스타일 ----------
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FILL_2 = PatternFill("solid", fgColor="4472C4")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

STATUS_FILL = {
    "correct":   PatternFill("solid", fgColor="C6EFCE"),
    "wrong":     PatternFill("solid", fgColor="FFC7CE"),
    "null_miss": PatternFill("solid", fgColor="FFEB9C"),
    "null_ok":   PatternFill("solid", fgColor="DDEBF7"),
}

FIELD_GROUPS = [
    # wishlist_count는 v3 평가 정책에서 제외됨 (GT 라벨 어렵고 follower 혼동 큼)
    ("텍스트추출 (가격·리뷰·상품명)", ["product_name", "original_price", "has_discount",
                          "discounted_price", "discount_rate",
                          "review_count", "review_score"], "ED7D31"),
    ("메타 (사진 유형·가시성)", ["shot_type", "visibility"], "A5A5A5"),
    ("스타일", ["style_keywords"], "70AD47"),
    ("마케팅 트리거", ["trend_hype", "bundle", "confidence"], "FFC000"),
    ("선택 필드 (배송·브랜드)", ["delivery_info", "delivery_fee", "brand"], "9DC3E6"),
]


def header(ws, row, cells, fill=HEADER_FILL):
    for c, v in enumerate(cells, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = WRAP_CENTER
        cell.border = BORDER


def auto_width(ws, widths):
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ---------------------------------------------------------------------------
# 시트 1. 요약
# ---------------------------------------------------------------------------

def sheet_summary(wb, summary):
    ws = wb.create_sheet("1. 요약")
    ws["A1"] = f"시험 0 v2 평가 — 매칭 {summary['n_images']}장"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    header(ws, 3, ["그룹", "필드", "정확도", "correct", "wrong", "null_miss"])
    row = 4
    per_field = summary["per_field"]
    for grp, fields, color in FIELD_GROUPS:
        for f in fields:
            m = per_field.get(f)
            if not m:
                continue
            acc = m.get("accuracy_excl_null_ok")
            ws.cell(row=row, column=1, value=grp).fill = PatternFill("solid", fgColor=color)
            ws.cell(row=row, column=2, value=f).alignment = LEFT_CENTER
            ws.cell(row=row, column=3, value=acc if acc is not None else None).number_format = "0.0%"
            ws.cell(row=row, column=4, value=m["correct"])
            ws.cell(row=row, column=5, value=m["wrong"])
            ws.cell(row=row, column=6, value=m["null_miss"])
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = BORDER
            row += 1

    # 색상 그라데이션 (정확도)
    last = row - 1
    if last >= 4:
        ws.conditional_formatting.add(
            f"C4:C{last}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                           mid_type="num", mid_value=0.7, mid_color="FFEB84",
                           end_type="num", end_value=1, end_color="63BE7B"),
        )
    auto_width(ws, [25, 22, 12, 10, 10, 12])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# 시트 2. 플랫폼별
# ---------------------------------------------------------------------------

def sheet_platform(wb, df):
    ws = wb.create_sheet("2. 플랫폼별")
    ws["A1"] = "플랫폼 × 필드 정확도 히트맵"
    ws["A1"].font = Font(bold=True, size=14)

    pivot = (df.assign(c=(df.status == "correct").astype(int),
                       e=(df.status != "null_ok").astype(int))
             .groupby(["platform", "field"]).agg(c=("c", "sum"), e=("e", "sum"))
             .reset_index())
    pivot["acc"] = pivot.apply(lambda r: r["c"] / r["e"] if r["e"] else None, axis=1)
    mat = pivot.pivot(index="platform", columns="field", values="acc")

    # 필드 순서: 그룹 정렬
    ordered_fields = [f for _, fs, _ in FIELD_GROUPS for f in fs if f in mat.columns]
    mat = mat[ordered_fields]

    # 헤더
    header(ws, 3, ["platform"] + ordered_fields)
    row = 4
    for plat in mat.index:
        ws.cell(row=row, column=1, value=plat).font = Font(bold=True)
        ws.cell(row=row, column=1).border = BORDER
        for c, f in enumerate(ordered_fields, start=2):
            v = mat.loc[plat, f]
            cell = ws.cell(row=row, column=c, value=v if pd.notna(v) else None)
            if pd.notna(v):
                cell.number_format = "0%"
            cell.border = BORDER
            cell.alignment = WRAP_CENTER
        row += 1

    last = row - 1
    last_col = get_column_letter(1 + len(ordered_fields))
    ws.conditional_formatting.add(
        f"B4:{last_col}{last}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.7, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"),
    )
    auto_width(ws, [12] + [13] * len(ordered_fields))
    ws.freeze_panes = "B4"


# ---------------------------------------------------------------------------
# 시트 3. 마케팅 P/R/F1
# ---------------------------------------------------------------------------

def sheet_marketing(wb, summary):
    ws = wb.create_sheet("3. 마케팅 트리거")
    ws["A1"] = "마케팅 트리거 P/R/F1 (positive = GT 1)"
    ws["A1"].font = Font(bold=True, size=14)

    header(ws, 3, ["축", "n", "Precision", "Recall", "F1"])
    row = 4
    for axis, m in summary.get("marketing_metrics", {}).items():
        ws.cell(row=row, column=1, value=axis).font = Font(bold=True)
        ws.cell(row=row, column=2, value=m["n"])
        for c, k in enumerate(["precision", "recall", "f1"], start=3):
            cell = ws.cell(row=row, column=c, value=m[k])
            cell.number_format = "0.00"
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = WRAP_CENTER
        row += 1

    # 색상 그라데이션 (P/R/F1)
    ws.conditional_formatting.add(
        f"C4:E{row-1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"),
    )

    ws[f"A{row+2}"] = "해석:"
    ws[f"A{row+2}"].font = Font(bold=True)
    notes = [
        "Precision 낮음 = 추출이 1로 잡았는데 GT는 0 (과탐지)",
        "Recall 낮음 = GT가 1인데 추출이 0으로 놓침",
        "v2 패턴: Recall=1.0 (다 잡음) but Precision <0.25 (false positive 많음)",
        "→ v3에서 trend_hype/bundle/confidence 키워드 화이트리스트로 좁힘 필요",
    ]
    for i, n in enumerate(notes):
        ws[f"A{row+3+i}"] = "  • " + n
    auto_width(ws, [16, 8, 12, 12, 12])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# 시트 4. 플래그
# ---------------------------------------------------------------------------

def sheet_flags(wb, summary):
    ws = wb.create_sheet("4. 플래그")
    ws["A1"] = "특수 케이스 플래그 (v3: wishlist 추출 안 함 → follower_confusion 사라짐)"
    ws["A1"].font = Font(bold=True, size=14)

    flags = summary.get("flags", {})
    cc = flags.get("coupon_confusion", []) or []

    row = 3
    ws.cell(row=row, column=1, value=f"coupon_confusion ({len(cc)}건) — discounted_price가 GT의 85% 미만").font = Font(bold=True, size=12)
    row += 1
    if cc:
        header(ws, row, ["image_id", "GT 할인가", "추출 할인가", "추정 원인"])
        row += 1
        for f in cc:
            ws.cell(row=row, column=1, value=f["image_id"])
            ws.cell(row=row, column=2, value=f["gt"]).number_format = "#,##0"
            ws.cell(row=row, column=3, value=f["pred"]).number_format = "#,##0"
            ws.cell(row=row, column=4, value="쿠폰가 혼동")
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="(없음)").font = Font(italic=True, color="808080")
        row += 1

    auto_width(ws, [25, 14, 14, 12, 22])


# ---------------------------------------------------------------------------
# 시트 5. 이미지별 상세 (image_id × 필드 grid, status 색상)
# ---------------------------------------------------------------------------

def sheet_per_image(wb, df):
    ws = wb.create_sheet("5. 이미지별 상세")
    ws["A1"] = "이미지 × 필드 status (색상: 초록=correct, 빨강=wrong, 노랑=null_miss, 파랑=null_ok)"
    ws["A1"].font = Font(bold=True, size=14)

    # pivot: image_id × field → status
    pv_status = df.pivot_table(index=["platform", "image_id"], columns="field",
                                values="status", aggfunc="first")
    fields = [f for _, fs, _ in FIELD_GROUPS for f in fs if f in pv_status.columns]
    pv_status = pv_status[fields]

    # 헤더
    header(ws, 3, ["platform", "image_id"] + fields)
    row = 4
    for (plat, iid), values in pv_status.iterrows():
        ws.cell(row=row, column=1, value=plat).border = BORDER
        ws.cell(row=row, column=2, value=iid).border = BORDER
        for c, f in enumerate(fields, start=3):
            status = values[f]
            cell = ws.cell(row=row, column=c, value=status if pd.notna(status) else "")
            cell.alignment = WRAP_CENTER
            cell.border = BORDER
            if pd.notna(status) and status in STATUS_FILL:
                cell.fill = STATUS_FILL[status]
        row += 1

    auto_width(ws, [10, 24] + [13] * len(fields))
    ws.freeze_panes = "C4"


# ---------------------------------------------------------------------------
# 시트 6. GT vs 추출 값 비교 (틀린 케이스만)
# ---------------------------------------------------------------------------

def load_jsonl(p):
    return [json.loads(l) for l in p.open(encoding="utf-8") if l.strip()]


def load_predictions(d):
    by_id = {}
    for p in d.glob("*.jsonl"):
        for row in load_jsonl(p):
            iid = row["image_id"]
            run = row.get("run", 1)
            existing = by_id.get(iid)
            if existing is None or run > existing[0]:
                by_id[iid] = (run, row.get("result") or {})
    return {k: v[1] for k, v in by_id.items()}


PRED_TO_GT = {}  # v3는 GT 양식과 1:1 일치


def _val(pred, key):
    if key in pred:
        v = pred[key]
        return v
    for pk, gk in PRED_TO_GT.items():
        if gk == key and pk in pred:
            v = pred[pk]
            if v is True: return 1
            if v is False: return 0
            return v
    return None


def sheet_wrong_cases(wb, df, gts, preds):
    ws = wb.create_sheet("6. 틀린 케이스")
    ws["A1"] = "wrong/null_miss 케이스만 — GT vs 추출 비교 (status로 필터)"
    ws["A1"].font = Font(bold=True, size=14)

    header(ws, 3, ["image_id", "platform", "field", "status", "GT 값", "추출 값", "detail"])
    row = 4
    gt_lookup = {g["image_id"]: g for g in gts}
    interesting = df[df["status"].isin(["wrong", "null_miss"])].sort_values(["field", "image_id"])
    for _, r in interesting.iterrows():
        gt = gt_lookup.get(r["image_id"], {})
        pred = preds.get(r["image_id"], {})
        ws.cell(row=row, column=1, value=r["image_id"])
        ws.cell(row=row, column=2, value=r["platform"])
        ws.cell(row=row, column=3, value=r["field"])
        ws.cell(row=row, column=4, value=r["status"]).fill = STATUS_FILL.get(r["status"], PatternFill())
        gt_v = gt.get(r["field"])
        pred_v = _val(pred, r["field"])
        ws.cell(row=row, column=5, value=str(gt_v)[:80] if gt_v is not None else "")
        ws.cell(row=row, column=6, value=str(pred_v)[:80] if pred_v is not None else "")
        ws.cell(row=row, column=7, value=r.get("detail") if pd.notna(r.get("detail")) else "")
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = BORDER
        row += 1

    auto_width(ws, [22, 10, 18, 11, 32, 32, 24])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{row-1}"


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    summary = json.loads((OUT_DIR / "summary.json").read_text(encoding="utf-8"))
    df = pd.read_parquet(OUT_DIR / "eval_results.parquet")
    gts = load_jsonl(GT_PATH)
    preds = load_predictions(PRED_DIR)

    wb = Workbook()
    wb.remove(wb.active)

    sheet_summary(wb, summary)
    sheet_platform(wb, df)
    sheet_marketing(wb, summary)
    sheet_flags(wb, summary)
    sheet_per_image(wb, df)
    sheet_wrong_cases(wb, df, gts, preds)

    out = OUT_DIR / "score_summary.xlsx"
    wb.save(out)
    print(f"[OK] {out}")


if __name__ == "__main__":
    main()
