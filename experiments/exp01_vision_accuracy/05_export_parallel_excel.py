"""GT 옆에 추출값을 나란히 보여주는 엑셀.

Outputs:
  outputs/gt_vs_pred.xlsx — 시트 2개:
    1. 병렬 비교 — 각 image_id 한 row, 필드별 [GT, 추출, status] 3열씩
    2. 차이 (틀린 케이스만) — 필드 단위 long form, 컬러 + auto_filter

색상:
  초록  = correct
  빨강  = wrong
  노랑  = null_miss (GT 있는데 추출 null)
  파랑  = null_ok   (둘 다 null)
"""

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
OUT_DIR = BASE / "outputs"
GT_PATH = BASE / "gt" / "gt_labels.jsonl"
PRED_DIR = BASE.parent / "exp00_vision_extraction" / "results" / "gemini_v3"

# v3는 GT 양식과 1:1 일치 → 매핑 불필요. 빈 dict로 둠.
PRED_TO_GT = {}

# 보여줄 필드 (그룹 순)
FIELDS = [
    ("product_name",     "ED7D31"),
    ("original_price",   "ED7D31"),
    ("has_discount",     "ED7D31"),
    ("discounted_price", "ED7D31"),
    ("discount_rate",    "ED7D31"),
    ("review_count",     "ED7D31"),
    ("review_score",     "ED7D31"),
    # wishlist_count는 v3 평가 정책에서 제외 — 보기용으로도 빼서 깔끔하게
    ("shot_type",        "A5A5A5"),
    ("visibility",       "A5A5A5"),
    ("style_keywords",   "70AD47"),
    ("trend_hype",       "FFC000"),
    ("bundle",           "FFC000"),
    ("confidence",       "FFC000"),
    ("marketing_phrases","FFC000"),
    ("delivery_info",    "9DC3E6"),
    ("delivery_fee",     "9DC3E6"),
    ("brand",            "9DC3E6"),
]

STATUS_FILL = {
    "correct":   PatternFill("solid", fgColor="C6EFCE"),
    "wrong":     PatternFill("solid", fgColor="FFC7CE"),
    "null_miss": PatternFill("solid", fgColor="FFEB9C"),
    "null_ok":   PatternFill("solid", fgColor="DDEBF7"),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
SUB_FONT = Font(bold=True, size=9, color="262626")


def load_jsonl(p):
    return [json.loads(l) for l in p.open(encoding="utf-8") if l.strip()]


def load_predictions(d):
    by_id = {}
    for p in d.glob("*.jsonl"):
        for row in load_jsonl(p):
            iid = row["image_id"]
            run = row.get("run", 1)
            existing = by_id.get(iid)
            if existing is None or run > existing[0]:
                by_id[iid] = (run, row.get("result") or {})
    return {k: v[1] for k, v in by_id.items()}


def get_pred_value(pred, key):
    """v2 결과에서 GT 키 기준 값 가져오기 (PRED_TO_GT 매핑 적용)."""
    if key in pred:
        return pred[key]
    for pk, gk in PRED_TO_GT.items():
        if gk == key and pk in pred:
            v = pred[pk]
            if v is True:
                return 1
            if v is False:
                return 0
            return v
    return None


def fmt(v):
    """보기 좋게 변환."""
    if v is None or v == "":
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v) if v else "[]"
    if isinstance(v, bool):
        return "1" if v else "0"
    return v


def cell_status(gt_v, pred_v, field):
    """간이 status (eval_results.parquet 매칭 못해도 셀 색상 정도는 결정)."""
    g_null = gt_v in (None, "", []) or (isinstance(gt_v, list) and not gt_v)
    p_null = pred_v in (None, "", []) or (isinstance(pred_v, list) and not pred_v)
    if g_null and p_null:
        return "null_ok"
    if g_null and not p_null:
        return "wrong"  # GT null인데 추출 채움
    if not g_null and p_null:
        return "null_miss"
    # 양쪽 다 값 있음 — 단순 비교 (정확한 비교는 evaluate에 위임, 여기선 셀 색상용)
    if isinstance(gt_v, list) or isinstance(pred_v, list):
        gs = set(gt_v or [])
        ps = set(pred_v or [])
        return "correct" if gs == ps else "wrong"
    if isinstance(gt_v, str) and isinstance(pred_v, str):
        g, p = gt_v.strip(), pred_v.strip()
        return "correct" if (g == p or g in p or p in g) else "wrong"
    try:
        if abs(float(gt_v) - float(pred_v)) < 0.01:
            return "correct"
    except (TypeError, ValueError):
        pass
    return "correct" if gt_v == pred_v else "wrong"


# ---------------------------------------------------------------------------
# 시트 1. 병렬 비교
# ---------------------------------------------------------------------------

def sheet_parallel(wb, gts, preds, eval_status_lookup):
    ws = wb.create_sheet("1. GT vs 추출 병렬")
    ws["A1"] = "GT 옆에 실제 추출값을 나란히 — 색상: 초록 correct, 빨강 wrong, 노랑 null_miss, 파랑 null_ok"
    ws["A1"].font = Font(bold=True, size=13)

    # 행 1: 그룹 헤더 (병합)
    # 행 2: 필드명 (한 필드 = 2열: GT / 추출)
    # 행 3+: 데이터

    # row 2 헤더 baseline cells
    ws.cell(row=2, column=1, value="image_id").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor="305496")
    ws.cell(row=2, column=2, value="platform").font = HEADER_FONT
    ws.cell(row=2, column=2).fill = PatternFill("solid", fgColor="305496")
    ws.cell(row=3, column=1, value="").fill = PatternFill("solid", fgColor="305496")
    ws.cell(row=3, column=2, value="").fill = PatternFill("solid", fgColor="305496")

    col = 3
    field_col_map = {}  # field -> (gt_col, pred_col)
    for field, color in FIELDS:
        # row 1: 필드 그룹 (병합)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = ws.cell(row=1, column=col, value=field)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        # row 2: GT / 추출 sub-header
        for off, label in enumerate(["GT", "추출"]):
            sc = ws.cell(row=2, column=col + off, value=label)
            sc.fill = PatternFill("solid", fgColor=color)
            sc.font = SUB_FONT
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border = BORDER
        field_col_map[field] = (col, col + 1)
        col += 2

    # 데이터 row 3+
    row = 3
    gt_lookup = {g["image_id"]: g for g in gts}
    image_ids = sorted(set(list(gt_lookup.keys()) + list(preds.keys())))
    for iid in image_ids:
        gt = gt_lookup.get(iid, {})
        pred = preds.get(iid, {})
        ws.cell(row=row, column=1, value=iid).border = BORDER
        platform = gt.get("platform") or ("기타" if not iid.startswith(("ably", "musinsa", "zigzag"))
                                          else {"ably": "에이블리", "musinsa": "무신사", "zigzag": "지그재그"}.get(iid.split("_")[0]))
        ws.cell(row=row, column=2, value=platform).border = BORDER

        for field, _ in FIELDS:
            gt_col, pred_col = field_col_map[field]
            gt_v = gt.get(field)
            pred_v = get_pred_value(pred, field) if pred else None
            # eval status가 있으면 사용, 없으면 간이 추정
            key = (iid, field)
            status = eval_status_lookup.get(key) or cell_status(gt_v, pred_v, field)
            fill = STATUS_FILL.get(status, PatternFill())

            gc = ws.cell(row=row, column=gt_col, value=fmt(gt_v))
            gc.fill = fill
            gc.border = BORDER
            gc.alignment = Alignment(vertical="center", wrap_text=True)
            pc = ws.cell(row=row, column=pred_col, value=fmt(pred_v))
            pc.fill = fill
            pc.border = BORDER
            pc.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    # 컬럼 너비
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    for field, _ in FIELDS:
        gt_col, pred_col = field_col_map[field]
        # 필드별 너비 차등
        if field in ("product_name", "marketing_phrases"):
            w = 28
        elif field in ("delivery_info", "delivery_fee"):
            w = 20
        elif field in ("brand", "shot_type", "visibility", "style_keywords"):
            w = 14
        else:
            w = 11
        ws.column_dimensions[get_column_letter(gt_col)].width = w
        ws.column_dimensions[get_column_letter(pred_col)].width = w

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(col-1)}{row-1}"


# ---------------------------------------------------------------------------
# 시트 2. 차이만
# ---------------------------------------------------------------------------

def sheet_diff_only(wb, gts, preds, eval_df):
    ws = wb.create_sheet("2. 차이만 (wrong·null_miss)")
    ws["A1"] = "틀린 케이스만 — auto_filter로 필드/플랫폼 필터링 가능"
    ws["A1"].font = Font(bold=True, size=13)

    headers = ["image_id", "platform", "field", "status", "GT", "추출", "detail"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    gt_lookup = {g["image_id"]: g for g in gts}
    diff = eval_df[eval_df["status"].isin(["wrong", "null_miss"])].sort_values(["field", "platform", "image_id"])

    row = 3
    for _, r in diff.iterrows():
        gt = gt_lookup.get(r["image_id"], {})
        pred = preds.get(r["image_id"], {})
        gt_v = gt.get(r["field"])
        pred_v = get_pred_value(pred, r["field"]) if pred else None
        ws.cell(row=row, column=1, value=r["image_id"])
        ws.cell(row=row, column=2, value=r["platform"])
        ws.cell(row=row, column=3, value=r["field"])
        st = ws.cell(row=row, column=4, value=r["status"])
        st.fill = STATUS_FILL.get(r["status"], PatternFill())
        st.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=fmt(gt_v))
        ws.cell(row=row, column=6, value=fmt(pred_v))
        ws.cell(row=row, column=7, value=r.get("detail") if pd.notna(r.get("detail")) else "")
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 24
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{row-1}"


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    gts = load_jsonl(GT_PATH)
    preds = load_predictions(PRED_DIR)

    # eval_results.parquet에서 정확한 status 가져오기
    eval_path = OUT_DIR / "eval_results.parquet"
    if eval_path.exists():
        eval_df = pd.read_parquet(eval_path)
        eval_status_lookup = {(r.image_id, r.field): r.status for r in eval_df.itertuples()}
    else:
        eval_df = pd.DataFrame(columns=["image_id", "platform", "field", "status", "detail"])
        eval_status_lookup = {}

    wb = Workbook()
    wb.remove(wb.active)
    sheet_parallel(wb, gts, preds, eval_status_lookup)
    sheet_diff_only(wb, gts, preds, eval_df)

    out = OUT_DIR / "gt_vs_pred.xlsx"
    wb.save(out)
    print(f"[OK] {out}")


if __name__ == "__main__":
    main()
